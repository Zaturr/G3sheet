# -*- coding: utf-8 -*-
"""Genera CXP_Control_Semanal.xlsx: diseño simple para el usuario, lógica en tablas."""

from collections import defaultdict
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = "CXP_Control_Semanal.xlsx"
OUT_ALT = "CXP_Control_Semanal_generado.xlsx"

# Encabezados de tabla (sin espacios raros: referencias estructuradas)
HDR_MASTER = ["IdEdificio", "Prioridad", "Coordinador", "Estado"]
HDR_REG = ["Fecha", "Nombre", "IdEdificio", "Actividad", "Avance"]
HDR_CATALOGO = ["IdEdificio", "NombreEdificio", "Activo"]

MASTER_ROWS = [
    ["2566", "Facturación", ""],
    ["2587", "Facturación", ""],
    ["2573", "Facturación", ""],
    ["2616", "Rendición", ""],
    ["2586", "Rendición", ""],
    ["2614", "Rendición", ""],
    ["2575", "Rendición", ""],
    ["2604", "Rendición", ""],
]

REGISTRO_ROWS = [
    ["17/04/2026", "Yoiner", "2566", "Facturas JDC", "Completado"],
    ["17/04/2026", "Yoiner", "2587", "Paquete contador", "En curso"],
    ["17/04/2026", "Yoiner", "2614", "Rendición 2614", "Completado"],
    ["17/04/2026", "Genesis Quintero", "2616", "Revisión 2616 con MH", "Completado"],
    ["17/04/2026", "Genesis Quintero", "2586", "Rendición 2586", "Completado"],
    ["17/04/2026", "Genesis Quintero", "2575", "Rendición Campo Neblina", "En curso"],
    ["17/04/2026", "Yumilka", "2604", "Corpoelec", "En curso"],
    ["17/04/2026", "Yumilka", "2573", "Archivos", "En curso"],
    ["17/04/2026", "Yumilka", "2587", "Pagos proveedores", "En curso"],
    ["17/04/2026", "Sandra", "2616", "Comunicación JDC 2616", "Completado"],
    ["17/04/2026", "Sandra", "2604", "Cuadro CANTV", "En curso"],
    ["17/04/2026", "Sandra", "2586", "Estados de cuenta", "En curso"],
    ["17/04/2026", "Genesis Quintero", "2614", "Pagos", "En curso"],
    ["17/04/2026", "Yoiner", "2573", "Conciliación 8667", "En curso"],
]

CATALOGO_ROWS = [
    ["2566", "", "Si"],
    ["2573", "", "Si"],
    ["2575", "", "Si"],
    ["2586", "", "Si"],
    ["2587", "", "Si"],
    ["2604", "", "Si"],
    ["2614", "", "Si"],
    ["2616", "", "Si"],
]

EMPLEADOS_LIST = ["Yoiner", "Genesis Quintero", "Yumilka", "Sandra"]
EMPLEADOS = ",".join(EMPLEADOS_LIST)
AVANCES = "En curso,Completado"

FORMULA_ESTADO = (
    '=IF(COUNTIFS(\'2 Registro diario\'!$C:$C,A{row},'
    '\'2 Registro diario\'!$E:$E,"Completado")>0,'
    '"FINALIZADO","PENDIENTE")'
)

# Colores (hex sin #)
C_HEADER = "2F5597"
C_TITLE_BG = "D6E3F4"
C_TITLE_FG = "1F3864"
C_SUBTLE = "F2F2F2"
C_TEXT = "333333"


def _thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def build_inicio(ws):
    ws.title = "1 Inicio"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 14

    title = ws.cell(2, 2, "Control semanal CXP")
    title.font = Font(size=22, bold=True, color=C_TITLE_FG)
    title.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)

    sub = ws.cell(3, 2, "Empieza aquí: tres pestañas, en orden.")
    sub.font = Font(size=12, italic=True, color="666666")
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)

    bloques = [
        (
            "Qué haces tú normalmente",
            "1) Abre la pestaña «2 Registro diario».\n\n"
            "2) Cada vez que hagas algo, añade una fila nueva.\n"
            "   • Elige tu nombre y escribe el identificador del edificio.\n"
            "   • Escribe con frases simples qué hiciste.\n"
            "   • Pon «Completado» solo si esa parte ya quedó lista del todo.\n\n"
            "3) Puedes tener varias filas el mismo día: una por cada tema distinto.",
        ),
        (
            "Qué hace el equipo cada lunes",
            "• Actualizar la pestaña «3 Plan semanal» con los edificios de la semana.\n\n"
            "• Vaciar el registro del día (borrar las filas de datos, no los encabezados azules).",
        ),
        (
            "Cómo leer el «Estado» en el plan",
            "• PENDIENTE = aún nadie marcó esa tarea como terminada en el registro.\n\n"
            "• FINALIZADO = al menos una vez alguien puso «Completado» para ese edificio.\n\n"
            "Tú no escribas en esa columna: se calcula sola.",
        ),
    ]

    r = 6
    for titulo, texto in bloques:
        cell_t = ws.cell(r, 2, titulo)
        cell_t.font = Font(size=13, bold=True, color=C_TITLE_FG)
        cell_t.fill = PatternFill("solid", fgColor=C_TITLE_BG)
        cell_t.border = _thin_border()
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        r += 1
        cell_b = ws.cell(r, 2, texto)
        cell_b.font = Font(size=11, color=C_TEXT)
        cell_b.alignment = Alignment(wrap_text=True, vertical="top")
        cell_b.fill = PatternFill("solid", fgColor="FFFFFF")
        cell_b.border = _thin_border()
        ws.merge_cells(start_row=r, start_column=2, end_row=r + 3, end_column=7)
        ws.row_dimensions[r].height = 28
        for extra in range(1, 4):
            ws.row_dimensions[r + extra].height = 18
        r += 5

    pie = ws.cell(r + 1, 2, "¿Dudas? Pide ayuda a quien cargue el plan semanal — el archivo no pierde fórmulas al borrar filas del registro.")
    pie.font = Font(size=10, italic=True, color="888888")
    ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=7)


def apply_banner(ws, row_start, main_text, hint_text, merge_cols=5):
    ws.merge_cells(
        start_row=row_start,
        start_column=1,
        end_row=row_start,
        end_column=merge_cols,
    )
    m = ws.cell(row_start, 1, main_text)
    m.font = Font(size=14, bold=True, color="FFFFFF")
    m.fill = PatternFill("solid", fgColor=C_HEADER)
    m.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row_start].height = 32

    ws.merge_cells(
        start_row=row_start + 1,
        start_column=1,
        end_row=row_start + 1,
        end_column=merge_cols,
    )
    h = ws.cell(row_start + 1, 1, hint_text)
    h.font = Font(size=10, color="444444")
    h.fill = PatternFill("solid", fgColor=C_SUBTLE)
    h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row_start + 1].height = 36
    ws.row_dimensions[row_start + 2].height = 8


def style_header_row(ws, row, n_cols):
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor=C_HEADER)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row, c)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_resumen_por_empleado(ws, empleados):
    ws.title = "4 Vista por empleado"
    ws.sheet_view.showGridLines = False

    apply_banner(
        ws,
        1,
        "Vista rápida por empleado",
        "Cada columna se actualiza sola con el Registro diario para esa persona.",
        max(1, len(empleados) * 2),
    )

    start_row = 5
    first_data_row = start_row + 1
    for idx, empleado in enumerate(empleados):
        col = 1 + idx * 2
        ws.column_dimensions[get_column_letter(col)].width = 42
        ws.column_dimensions[get_column_letter(col + 1)].width = 2

        h = ws.cell(start_row, col, empleado)
        h.font = Font(bold=True, color="FFFFFF", size=11)
        h.fill = PatternFill("solid", fgColor=C_HEADER)
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        h.border = _thin_border()

        # Formula tipo Google Sheets: se expande sola al agregar filas.
        col_l = get_column_letter(col)
        formula = (
            '=IFNA(ARRAYFORMULA('
            'FILTER(\'2 Registro diario\'!$A$6:$A1000; \'2 Registro diario\'!$B$6:$B1000='
            f'{col_l}$5) & " | " & '
            'FILTER(\'2 Registro diario\'!$C$6:$C1000; \'2 Registro diario\'!$B$6:$B1000='
            f'{col_l}$5) & " | " & '
            'FILTER(\'2 Registro diario\'!$D$6:$D1000; \'2 Registro diario\'!$B$6:$B1000='
            f'{col_l}$5) & " | " & '
            'FILTER(\'2 Registro diario\'!$E$6:$E1000; \'2 Registro diario\'!$B$6:$B1000='
            f'{col_l}$5)); "(sin registros)")'
        )
        c = ws.cell(first_data_row, col, formula)
        c.font = Font(size=10, color=C_TEXT)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = _thin_border()


def build_catalogo_edificios(ws, catalogo_rows):
    ws.title = "8 Catalogo edificios"
    ws.sheet_view.showGridLines = False
    apply_banner(
        ws,
        1,
        "Catalogo de edificios",
        "Agrega aqui todos los edificios (actuales y futuros). Este catalogo alimenta el conteo semanal.",
        3,
    )

    top = 5
    for c, h in enumerate(HDR_CATALOGO, 1):
        ws.cell(top, c, h)
    style_header_row(ws, top, 3)

    for i, row in enumerate(catalogo_rows, top + 1):
        for c, val in enumerate(row, 1):
            ws.cell(i, c, val)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 12
    ws.freeze_panes = f"A{top + 1}"

    ref = f"A{top}:C{top + len(catalogo_rows)}"
    tbl = Table(displayName="Catalogo_Edificios", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4", showRowStripes=True, showFirstColumn=False
    )
    ws.add_table(tbl)

    dv_activo = DataValidation(type="list", formula1='"Si,No"', allow_blank=True)
    ws.add_data_validation(dv_activo)
    dv_activo.add(f"C{top + 1}:C400")


def build_reporte_semanal(ws, empleados):
    ws.title = "5 Reporte semanal"
    ws.sheet_view.showGridLines = False

    apply_banner(
        ws,
        1,
        "Reporte semanal — mismo formato del registro diario",
        "Al final del dia registra en «2 Registro diario» y aqui se consolida automaticamente por semana.",
        5,
    )

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 34
    ws.column_dimensions["E"].width = 18

    ws.cell(5, 1, "Fecha inicio semana")
    ws.cell(5, 2, "14/04/2026")
    ws.cell(6, 1, "Fecha fin semana")
    ws.cell(6, 2, "20/04/2026")
    ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=5)
    ws.cell(8, 1, '="Semana de " & TEXT($B$5; "dd/mm/yyyy") & " a " & TEXT($B$6; "dd/mm/yyyy")')
    ws.cell(8, 1).font = Font(size=11, bold=True, color=C_TITLE_FG)
    ws.cell(8, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(8, 1).fill = PatternFill("solid", fgColor=C_TITLE_BG)
    ws.cell(8, 1).border = _thin_border()

    ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=5)
    ws.cell(9, 1, "Usa formato dd/mm/aaaa en B5 y B6.")
    ws.cell(9, 1).font = Font(size=9, italic=True, color="666666")
    ws.cell(9, 1).alignment = Alignment(horizontal="center", vertical="center")

    for rr in (5, 6):
        ws.cell(rr, 1).font = Font(bold=True, color=C_TITLE_FG)
        ws.cell(rr, 1).fill = PatternFill("solid", fgColor=C_TITLE_BG)
        ws.cell(rr, 1).border = _thin_border()
        ws.cell(rr, 2).border = _thin_border()
        ws.cell(rr, 2).number_format = "dd/mm/yyyy"
        ws.cell(rr, 2).alignment = Alignment(horizontal="center")

    top = 11
    for c, h in enumerate(HDR_REG, 1):
        ws.cell(top, c, h)
    style_header_row(ws, top, len(HDR_REG))
    ws.freeze_panes = "A12"

    ws.cell(
        12,
        1,
        '=IFNA(FILTER(\'2 Registro diario\'!$A$6:$A1000; (\'2 Registro diario\'!$A$6:$A1000 >= $B$5) * (\'2 Registro diario\'!$A$6:$A1000 <= $B$6)); "")',
    )
    ws.cell(
        12,
        2,
        '=IFNA(FILTER(\'2 Registro diario\'!$B$6:$B1000; (\'2 Registro diario\'!$A$6:$A1000 >= $B$5) * (\'2 Registro diario\'!$A$6:$A1000 <= $B$6)); "")',
    )
    ws.cell(
        12,
        3,
        '=IFNA(FILTER(\'2 Registro diario\'!$C$6:$C1000; (\'2 Registro diario\'!$A$6:$A1000 >= $B$5) * (\'2 Registro diario\'!$A$6:$A1000 <= $B$6)); "")',
    )
    ws.cell(
        12,
        4,
        '=IFNA(FILTER(\'2 Registro diario\'!$D$6:$D1000; (\'2 Registro diario\'!$A$6:$A1000 >= $B$5) * (\'2 Registro diario\'!$A$6:$A1000 <= $B$6)); "")',
    )
    ws.cell(
        12,
        5,
        '=IFNA(FILTER(\'2 Registro diario\'!$E$6:$E1000; (\'2 Registro diario\'!$A$6:$A1000 >= $B$5) * (\'2 Registro diario\'!$A$6:$A1000 <= $B$6)); "")',
    )


def _parse_fecha(fecha_str):
    return datetime.strptime(fecha_str, "%d/%m/%Y").date()


def _week_bounds(registro_rows):
    fechas = sorted(_parse_fecha(r[0]) for r in registro_rows)
    if not fechas:
        hoy = datetime.today().date()
        inicio = hoy - timedelta(days=hoy.weekday())
        return inicio, inicio + timedelta(days=6)
    inicio = fechas[0] - timedelta(days=fechas[0].weekday())
    fin = inicio + timedelta(days=6)
    return inicio, fin


def _render_text_sheet(ws, title, lines):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 140
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
    ws.cell(1, 1, title).font = Font(size=14, bold=True, color="FFFFFF")
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=C_HEADER)
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    for idx, line in enumerate(lines, 3):
        c = ws.cell(idx, 1, line)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(size=11, color=C_TEXT)


def build_reporte_diario_texto(ws, _registro_rows, _master_rows, empleados):
    ws.title = "6 Reporte diario texto"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 140

    ws.cell(1, 1, "Reporte diario en texto (dinámico)")
    ws.cell(1, 1).font = Font(size=14, bold=True, color="FFFFFF")
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=C_HEADER)
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.cell(2, 1, "Fecha objetivo")
    ws.cell(2, 2, '=IFERROR(IF(MAX(\'2 Registro diario\'!$A$6:$A)=0; TODAY(); MAX(\'2 Registro diario\'!$A$6:$A)); TODAY())')
    ws.cell(2, 2).number_format = "dd/mm/yyyy"
    ws.cell(2, 1).font = Font(bold=True, color=C_TITLE_FG)

    ws.cell(4, 1, '="*Resumen gestión cxp del " & TEXT($B$2; "dd/mm/yyyy") & ":*"')
    ws.cell(6, 1, '="*Pendientes del " & TEXT($B$2; "dd/mm/yyyy") & " según planificación:*"')
    ws.cell(7, 1, '=IFNA(ARRAYFORMULA("-Revisar gestión del edificio " & FILTER(\'3 Plan semanal\'!$A$6:$A; \'3 Plan semanal\'!$A$6:$A<>""; MAP(\'3 Plan semanal\'!$A$6:$A; LAMBDA(edificio; COUNTIFS(\'2 Registro diario\'!$C$6:$C; edificio; \'2 Registro diario\'!$A$6:$A; $B$2; \'2 Registro diario\'!$E$6:$E; "Completado")=0))) & "."); "-Sin pendientes según planificación.")')

    ws.cell(15, 1, '="*Realizados pendientes del " & TEXT($B$2; "dd/mm/yyyy") & ":*"')
    ws.cell(16, 1, '=IFNA(ARRAYFORMULA("-" & FILTER(\'2 Registro diario\'!$D$6:$D; \'2 Registro diario\'!$A$6:$A=$B$2; \'2 Registro diario\'!$E$6:$E="Completado") & " (Edificio " & FILTER(\'2 Registro diario\'!$C$6:$C; \'2 Registro diario\'!$A$6:$A=$B$2; \'2 Registro diario\'!$E$6:$E="Completado") & ")."); "-Sin pendientes realizados en esta fecha.")')

    ws.cell(24, 1, '="*Resumen individual gestión cxp del " & TEXT($B$2; "dd/mm/yyyy") & ":*"')

    row = 26
    for nombre in empleados:
        ws.cell(row, 1, f"{nombre.upper()}:")
        ws.cell(
            row + 1,
            1,
            '=IFNA(ARRAYFORMULA("-" & FILTER(\'2 Registro diario\'!$D$6:$D; \'2 Registro diario\'!$A$6:$A=$B$2; \'2 Registro diario\'!$B$6:$B="{n}") & " (Edificio " & FILTER(\'2 Registro diario\'!$C$6:$C; \'2 Registro diario\'!$A$6:$A=$B$2; \'2 Registro diario\'!$B$6:$B="{n}") & ")."); "-Sin registros en la fecha.")'.format(
                n=nombre
            ),
        )
        row += 7

    ws.cell(row, 1, '="Pendientes para el " & TEXT($B$2+3; "dd/mm/yyyy") & ":"')
    ws.cell(row + 1, 1, '=IFNA(ARRAYFORMULA("-Revisar gestión del edificio " & FILTER(\'3 Plan semanal\'!$A$6:$A; \'3 Plan semanal\'!$A$6:$A<>""; MAP(\'3 Plan semanal\'!$A$6:$A; LAMBDA(edificio; COUNTIFS(\'2 Registro diario\'!$C$6:$C; edificio; \'2 Registro diario\'!$A$6:$A; $B$2; \'2 Registro diario\'!$E$6:$E; "Completado")=0)))); "-Sin pendientes.")')

    for r in range(1, row + 3):
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
        if r >= 4:
            ws.cell(r, 1).font = Font(size=11, color=C_TEXT)


def build_reporte_semanal_texto(ws, _registro_rows, _master_rows):
    ws.title = "7 Reporte semanal texto"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 140

    ws.cell(1, 1, "Reporte semanal en texto (dinámico)")
    ws.cell(1, 1).font = Font(size=14, bold=True, color="FFFFFF")
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=C_HEADER)
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.cell(2, 1, "Fecha inicio semana")
    ws.cell(2, 2, "14/04/2026")
    ws.cell(3, 1, "Fecha fin semana")
    ws.cell(3, 2, "20/04/2026")
    ws.cell(2, 1).font = Font(bold=True, color=C_TITLE_FG)
    ws.cell(3, 1).font = Font(bold=True, color=C_TITLE_FG)
    ws.cell(2, 2).number_format = "dd/mm/yyyy"
    ws.cell(3, 2).number_format = "dd/mm/yyyy"

    ws.cell(5, 1, '="*Resumen semana Laboral CXP (" & TEXT($B$2; "dd/mm/yyyy") & " al " & TEXT($B$3; "dd/mm/yyyy") & "):*"')
    ws.cell(7, 1, "*Prioridades:*")
    ws.cell(8, 1, '=IFNA(ARRAYFORMULA("-" & FILTER(\'3 Plan semanal\'!$B$6:$B1000; \'3 Plan semanal\'!$A$6:$A1000<>"") & ": " & FILTER(\'3 Plan semanal\'!$A$6:$A1000; \'3 Plan semanal\'!$A$6:$A1000<>"")); "-Sin prioridades registradas.")')

    ws.cell(18, 1, "*Logrado:*")
    ws.cell(19, 1, '=IFNA(ARRAYFORMULA("-" & FILTER(\'2 Registro diario\'!$D$6:$D1000; \'2 Registro diario\'!$A$6:$A1000>=$B$2; \'2 Registro diario\'!$A$6:$A1000<=$B$3; \'2 Registro diario\'!$E$6:$E1000="Completado") & " (Edificio " & FILTER(\'2 Registro diario\'!$C$6:$C1000; \'2 Registro diario\'!$A$6:$A1000>=$B$2; \'2 Registro diario\'!$A$6:$A1000<=$B$3; \'2 Registro diario\'!$E$6:$E1000="Completado") & " - " & TEXT(FILTER(\'2 Registro diario\'!$A$6:$A1000; \'2 Registro diario\'!$A$6:$A1000>=$B$2; \'2 Registro diario\'!$A$6:$A1000<=$B$3; \'2 Registro diario\'!$E$6:$E1000="Completado"); "dd/mm/yyyy") & ")."); "-No hay actividades marcadas como completadas en la semana.")')

    for r in range(1, 90):
        ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
        if r >= 5:
            ws.cell(r, 1).font = Font(size=11, color=C_TEXT)


def main():
    wb = Workbook()
    ws_i = wb.active
    build_inicio(ws_i)

    ws_r = wb.create_sheet("2 Registro diario", 1)
    ws_m = wb.create_sheet("3 Plan semanal", 2)
    ws_v = wb.create_sheet("4 Vista por empleado", 3)
    ws_rs = wb.create_sheet("5 Reporte semanal", 4)
    ws_rt_d = wb.create_sheet("6 Reporte diario texto", 5)
    ws_rt_s = wb.create_sheet("7 Reporte semanal texto", 6)
    ws_c = wb.create_sheet("8 Catalogo edificios", 7)

    # --- Registro: filas superiores + tabla ---
    apply_banner(
        ws_r,
        1,
        "Registro diario — escribe aquí cada actividad",
        "Una fila = una cosa que hiciste en un edificio. Usa listas solo para nombre y avance.",
        5,
    )
    ws_r.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5)
    tip_r = ws_r.cell(4, 1, "Consejo: al terminar la fila, pulsa Tab o Enter; puedes añadir filas desde el borde inferior de la tabla.")
    tip_r.font = Font(size=9, italic=True, color="666666")
    tip_r.alignment = Alignment(horizontal="center", vertical="center")
    ws_r.row_dimensions[4].height = 20
    tbl_top_r = 5
    for c, h in enumerate(HDR_REG, 1):
        ws_r.cell(tbl_top_r, c, h)
    for i, row in enumerate(REGISTRO_ROWS, tbl_top_r + 1):
        for c, val in enumerate(row, 1):
            ws_r.cell(i, c, val)

    n_data_r = len(REGISTRO_ROWS)
    last_r = tbl_top_r + n_data_r
    ref_r = f"A{tbl_top_r}:E{last_r}"

    for c in range(1, 6):
        ws_r.column_dimensions[get_column_letter(c)].width = 26 if c == 4 else 18

    style_header_row(ws_r, tbl_top_r, 5)
    ws_r.freeze_panes = f"A{tbl_top_r + 1}"

    tbl_reg = Table(displayName="Registro", ref=ref_r)
    tbl_reg.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True, showFirstColumn=False
    )
    ws_r.add_table(tbl_reg)

    # Validaciones (Nombre, IdEdificio, Avance)
    dv_nombre = DataValidation(
        type="list",
        formula1=f'"{EMPLEADOS}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Nombre",
        error="Elige un nombre de la lista.",
    )
    dv_avance = DataValidation(
        type="list",
        formula1=f'"{AVANCES}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Avance",
        error='Elige «En curso» o «Completado».',
    )
    ids_catalogo = ",".join(row[0] for row in CATALOGO_ROWS if row[0])
    dv_id_edificio = DataValidation(
        type="list",
        formula1=f'"{ids_catalogo}"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="IdEdificio",
        error="Elige un IdEdificio del catálogo.",
    )
    ws_r.add_data_validation(dv_nombre)
    ws_r.add_data_validation(dv_id_edificio)
    ws_r.add_data_validation(dv_avance)
    # Columnas B, C y E en datos (no encabezado)
    dv_nombre.add(f"B{tbl_top_r + 1}:B300")
    dv_id_edificio.add(f"C{tbl_top_r + 1}:C300")
    dv_avance.add(f"E{tbl_top_r + 1}:E300")

    # (Sin formato condicional: openpyxl a veces genera XML que Excel marca como
    # "reparar". Los colores de cabecera y tablas siguen siendo claros.)

    # --- Plan semanal ---
    apply_banner(
        ws_m,
        1,
        "Plan de la semana — edificios y prioridades",
        "Solo quien coordina edita aquí los edificios y prioridades. La columna Estado se calcula sola.",
        4,
    )
    tbl_top_m = 5
    for c, h in enumerate(HDR_MASTER, 1):
        ws_m.cell(tbl_top_m, c, h)
    for i, row in enumerate(MASTER_ROWS, tbl_top_m + 1):
        for c, val in enumerate(row, 1):
            ws_m.cell(i, c, val)

    n_m = len(MASTER_ROWS)
    last_m = tbl_top_m + n_m
    ref_m = f"A{tbl_top_m}:D{last_m}"

    for c in range(1, 5):
        ws_m.column_dimensions[get_column_letter(c)].width = 22 if c == 2 else 16

    style_header_row(ws_m, tbl_top_m, 4)
    ws_m.freeze_panes = f"A{tbl_top_m + 1}"

    # Tabla primero; luego fórmulas (mejor compatibilidad con referencias estructuradas)
    tbl_m = Table(displayName="Planificacion_Semanal", ref=ref_m)
    tbl_m.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False
    )
    ws_m.add_table(tbl_m)

    for r in range(tbl_top_m + 1, last_m + 1):
        ws_m.cell(r, 4, FORMULA_ESTADO.format(row=r))

    build_resumen_por_empleado(ws_v, EMPLEADOS_LIST)
    build_reporte_semanal(ws_rs, EMPLEADOS_LIST)
    build_reporte_diario_texto(ws_rt_d, REGISTRO_ROWS, MASTER_ROWS, EMPLEADOS_LIST)
    build_reporte_semanal_texto(ws_rt_s, REGISTRO_ROWS, MASTER_ROWS)
    build_catalogo_edificios(ws_c, CATALOGO_ROWS)

    try:
        wb.save(OUT)
        print(f"OK: {OUT}")
    except PermissionError:
        wb.save(OUT_ALT)
        print(
            f"OK: guardado como {OUT_ALT} (cierra {OUT} en Excel y vuelve a ejecutar "
            "el script para sobrescribir el archivo principal)."
        )


if __name__ == "__main__":
    main()
